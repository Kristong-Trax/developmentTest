# -*- coding: utf-8 -*-
import glob
import os
# from Trax.Cloud.Services.Connector.Logger import LoggerInitializer
import pandas as pd
from openpyxl import load_workbook
from KPIUtils.Utils.Validators.Template.TemplateValidator import TemplateValidator


def templateValidator():
    # pay attention if there any number column in template like product en please put string in the last column
    project_name = 'ccru'
    helper_path = '~/dev/trax_ace_factory/Projects/CCRU/Data/helper.xlsx'
    your_validate_path = "/home/TRAX/ilays/Desktop/Validation/"

    # file_path = '~/dev/trax_ace_factory/Projects/CCRU/Data/BFSuper.xlsx'
    # validate_file_path = '/home/TRAX/ilays/Desktop/1.xlsx'
    # validation_object = TemplateValidator(project_name, helper_path, file_path, validate_file_path)
    # validation_object.run()

    list_Of_Ignores = ["helper.xlsx", "KPIConvesion.xlsx", "gaps_guide.xlsx", "MT Shelf facings_2017.xlsx"]

    file_pre_path = '/home/TRAX/ilays/dev/trax_ace_factory/Projects/CCRU/Data/'

    # Define helper file
    wb = load_workbook(filename='helper.xlsx')
    workSheet = wb.worksheets[1]
    workSheet['B2'] = 'Type'
    workSheet['C2'] = 'Values'
    list_of_files = []

    # Iterate the files in the folder
    for filename in os.listdir(file_pre_path):
        if filename.endswith(".xlsx") and filename not in list_Of_Ignores:
            workBookFile = pd.ExcelFile(filename)
            sheetNames = workBookFile.sheet_names[0]
            print sheetNames
            workSheet['A2'] = sheetNames
            wb.save('helper.xlsx')
            file_path = '~/dev/trax_ace_factory/Projects/CCRU/Data/'+ filename
            print file_path
            list_of_files.append(file_path)
            validate_file_path = your_validate_path + filename
            validation_object = TemplateValidator(project_name, helper_path, file_path, validate_file_path)
            validation_object.run()

            #delete all sheets that are unrelated to the report
            workbook = load_workbook(validate_file_path)
            for s in workbook.get_sheet_names():
                sheet = workbook.get_sheet_by_name(s)
                if (len(sheet.columns) < 2):
                    workbook.remove_sheet(sheet)
                    workbook.save(validate_file_path)

    # combining all files in folder
    writer = pd.ExcelWriter(your_validate_path + "Final_report.xlsx", engine='xlsxwriter')
    dataFrame = pd.DataFrame()
    for f in glob.glob(your_validate_path + "*.xlsx"):
        df = pd.read_excel(f)
        dirPath, file_end = os.path.split(f)
        
        df.to_excel(writer, file_end[:30])

    writer.save()

    print "Finish iterate all files, comibed them to Final_report.xlsx"

# if __name__ == '__main__':
#     LoggerInitializer.init('start template validate')
#     templateValidator()
